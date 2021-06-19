# -*- coding: UTF-8 -*-

import os
import sys
import json
import openpyxl


def is_config_file_valid():
    if not "export_path" in config_json_data:
        print("config hasn't export_path")
        return False
    if not "server_code_path" in config_json_data:
        print("config hasn't server_code_path")
        return False
    if not "client_code_path" in config_json_data:
        print("config hasn't client_code_path")
        return False
    if not "namespace" in config_json_data:
        print("config hasn't namespace")
        return False
    return True


def is_config_valid(file_id):
    if not is_file_valid(file_id):
        print("config hasn't " + file_id)
        return False
    if not "upper_case" in config_json_data[file_id]:
        print(file_id + " hasn't upper_case")
        return False
    if not "lower_case" in config_json_data[file_id]:
        print(file_id + " hasn't lower_case")
        return False
    if not "sheet_list" in config_json_data[file_id]:
        print(file_id + " hasn't sheet_list")
        return False

    sheet_list = config_json_data[file_id]["sheet_list"]
    sheet_index = 0
    for sheet_info in sheet_list:
        if not "sheet_name" in sheet_info:
            print(file_id + "sheet_info " + sheet_index+" hasn't sheet_name")
            return False
        if not "data_id_row" in sheet_info:
            print(file_id + "sheet_info " + sheet_index+" hasn't data_id_row")
            return False
        if not "data_id_col" in sheet_info:
            print(file_id + "sheet_info " + sheet_index+" hasn't data_id_col")
            return False
        if not "export_type" in sheet_info:
            print(file_id + "sheet_info " + sheet_index+" hasn't export_type")
            return False
        sheet_index += 1
    return True


def is_file_valid(file_id):
    return file_id in config_json_data


def is_data_false(data):
    return data == "" or data == "0" or data == "False" or data == "false"


def get_enum_value(enum_type, enum_value):
    return enum_value_data[enum_type][enum_value]


def get_enum_type_comment(enum_type):
    if("#" in enum_json_data[enum_type]):
        return enum_json_data[enum_type]["#"]
    return ""


def get_enum_comment(enum_type, enum_value):
    if("#" in enum_json_data[enum_type][enum_value]):
        return enum_json_data[enum_type][enum_value]["#"]
    return ""


def get_server_data_type(data_type):
    if data_type == "bool":
        return "bool"
    if data_type == "int32":
        return "int32"
    if data_type == "float":
        return "float"
    if data_type == "string":
        return "std::string"
    if data_type == "bit":
        return "int32"
    if data_type == "vector<bool>":
        return "std::vector<bool>"
    if data_type == "vector<int32>":
        return "std::vector<int32>"
    if data_type == "vector<float>":
        return "std::vector<float>"
    if data_type == "vector<string>":
        return "std::vector<std::string>"
    if data_type in enum_json_data:
        return data_type
    else:
        data_type = data_type.lstrip("vector<").rstrip(">")
        if data_type in enum_json_data:
            return "std::vector<"+data_type+">"
    return ''


def get_client_data_type(data_type):
    if data_type == "bool":
        return "bool"
    if data_type == "int32":
        return "int32"
    if data_type == "float":
        return "float"
    if data_type == "string":
        return "FString"
    if data_type == "bit":
        return "int32"
    if data_type == "vector<bool>":
        return "TArray<bool>"
    if data_type == "vector<int32>":
        return "TArray<int32>"
    if data_type == "vector<float>":
        return "TArray<float>"
    if data_type == "vector<string>":
        return "TArray<FString>"
    if data_type in enum_json_data:
        return "TEnumAsByte<"+data_type+">"
    else:
        data_type = data_type.lstrip("vector<").rstrip(">")
        if data_type in enum_json_data:
            return "TArray<TEnumAsByte<"+data_type+">>"
    return ''


def get_parse_function_name(data_type):
    if data_type == "bool":
        return "ParseBool"
    if data_type == "int32":
        return "ParseInt"
    if data_type == "float":
        return "ParseFloat"
    if data_type == "string":
        return "ParseString"
    if data_type == "bit":
        return "ParseInt"
    if data_type == "vector<bool>":
        return "ParseVectorBool"
    if data_type == "vector<int32>":
        return "ParseVectorInt"
    if data_type == "vector<float>":
        return "ParseVectorFloat"
    if data_type == "vector<string>":
        return "ParseVectorString"
    if data_type in enum_json_data:
        return "ParseEnum<"+data_type+">"
    else:
        data_type = data_type.lstrip("vector<").rstrip(">")
        if data_type in enum_json_data:
            return "ParseVectorEnum<"+data_type+">"
    return ''


def parse_data(data_type, data):
    if data_type == "bool":
        data = str(data)
        data = data.strip()
        if is_data_false(data):
            return "0"
        return "1"
    if data_type == "int32":
        data = str(data)
        data = data.strip()
        if data == "":
            return "0"
        return str(int(float(data)))
    if data_type == "float":
        data = str(data)
        data = data.strip()
        if data == "":
            return "0"
        return str(float(data)).rstrip('0').rstrip('.')
    if data_type == "string":
        data = str(data)
        return data.replace("\"", "\"\"")
    if data_type == "bit":
        data = str(data)
        word_list = data.split(",")
        value = 0
        for word in word_list:
            value = value << 1 | int(parse_data("bool", word))
        return str(value)
    if data_type == "vector<bool>":
        data = str(data)
        word_list = data.split(",")
        data = "("
        is_first = True
        for word in word_list:
            if is_first:
                is_first = False
            else:
                data = data + ","
            data = data + parse_data("bool", word)
        data = data + ")"
        return data
    if data_type == "vector<int32>":
        data = str(data)
        word_list = data.split(",")
        data = "("
        is_first = True
        for word in word_list:
            if is_first:
                is_first = False
            else:
                data = data + ","
            if word == "":
                data = data + "0"
            else:
                data = data + parse_data("int32", word)
        data = data + ")"
        return data
    if data_type == "vector<float>":
        data = str(data)
        word_list = data.split(",")
        data = "("
        is_first = True
        for word in word_list:
            if is_first:
                is_first = False
            else:
                data = data + ","
            data = data + parse_data("float", word)
        data = data + ")"
        return data
    if data_type == "vector<string>":
        data = str(data)
        word_list = data.split(",")
        data = "("
        is_first = True
        for word in word_list:
            if is_first:
                is_first = False
            else:
                data = data + ","
            data = data + parse_data("string", word)
        data = data + ")"
        return data
    elif data_type in enum_json_data:
        data = str(data)
        data = data.strip()
        data = str(get_enum_value(data_type, data))
        return data
    else:
        data_type = data_type.lstrip("vector<").rstrip(">")
        if data_type in enum_json_data:
            data = str(data)
            word_list = data.split(",")
            data = "("
            is_first = True
            for word in word_list:
                if is_first:
                    is_first = False
                else:
                    data = data + ","
                data = data + parse_data(data_type, word)
            data = data + ")"
            return data

    return ''


def export_data(file_id):
    excel_file = openpyxl.load_workbook(file_path + file_dict[file_id])
    file_config = config_json_data[file_id]
    sheet_list = file_config["sheet_list"]

    is_first_sheet = True
    with open(file_path + config_json_data['export_path'] + file_config["upper_case"] + ".csv", 'w', encoding='utf-8') as export_file:
        for sheet_info in sheet_list:

            excel_sheet = excel_file[sheet_info["sheet_name"]]

            data_id_row = sheet_info["data_id_row"] - 1
            data_id_col = sheet_info["data_id_col"] - 1

            data_id_value = excel_sheet.cell(row=data_id_row+1, column=data_id_col+1).value

            if data_id_value != "DataId":
                print("DataId row or col not valid!")
                return

            excel_sheet_rows_data = tuple(excel_sheet.rows)
            excel_sheet_cols_data = tuple(excel_sheet.columns)

            data_id_row_value_list = excel_sheet_rows_data[data_id_row]
            data_id_col_value_list = excel_sheet_cols_data[data_id_col]

            data_row_max = len(data_id_col_value_list)
            data_col_max = len(data_id_row_value_list)

            if sheet_info["export_type"] == "Horizontal":

                csv_head_list = []
                for cell in excel_sheet_rows_data[data_id_row][data_id_col: data_col_max]:
                    if cell.value in [None, ""]:
                        break
                    csv_head_list.append(cell.value)

                csv_data_id_list = []
                for cell in excel_sheet_cols_data[data_id_col][data_id_row + 1: data_row_max]:
                    if cell.value in [None, ""]:
                        break
                    csv_data_id_list.append(cell.value)

                data_type_list = []
                for cell in excel_sheet_rows_data[data_id_row - 1][data_id_col: data_col_max]:
                    if cell.value in [None, ""]:
                        break
                    data_type_list.append(cell.value)

                data_col_count = len(csv_head_list)
                data_row_count = len(csv_data_id_list)

                if is_first_sheet:
                    is_first_sheet = False
                    export_file.write("---")
                    for csv_head in csv_head_list:
                        export_file.write(",")
                        export_file.write(csv_head)
                    export_file.write("\n")

                for row_index in range(0, data_row_count):
                    
                    data_type = data_type_list[0]
                    data_value = excel_sheet_rows_data[data_id_row + 1 + row_index][data_id_col].value
                    export_file.write(parse_data(data_type, data_value))

                    for col_index in range(0, data_col_count):

                        data_type = data_type_list[col_index]
                        data_value = excel_sheet_rows_data[data_id_row + 1 + row_index][data_id_col + col_index].value

                        export_file.write(",")
                        export_file.write("\"")
                        export_file.write(parse_data(data_type, data_value))
                        export_file.write("\"")
                    export_file.write("\n")

            elif sheet_info["export_type"] == "Vertical":

                csv_head_list = []
                for cell in excel_sheet_cols_data[data_id_col][data_id_row: data_row_max]:
                    if cell.value in [None, ""]:
                        break
                    csv_head_list.append(cell.value)

                csv_data_id_list = []
                for cell in excel_sheet_rows_data[data_id_row][data_id_col + 1: data_col_max]:
                    if cell.value in [None, ""]:
                        break
                    csv_data_id_list.append(cell.value)

                data_type_list = []
                for cell in excel_sheet_cols_data[data_id_col - 1][data_id_row: data_row_max]:
                    if cell.value in [None, ""]:
                        break
                    data_type_list.append(cell.value)

                data_row_count = len(csv_head_list)
                data_col_count = len(csv_data_id_list)

                if is_first_sheet:
                    is_first_sheet = False
                    export_file.write("---")
                    for csv_head in csv_head_list:
                        export_file.write(",")
                        export_file.write(csv_head)
                    export_file.write("\n")

                for col_index in range(0, data_col_count):

                    data_type = data_type_list[0]
                    data_value = excel_sheet_cols_data[data_id_col + 1 + col_index][data_id_row].value
                    export_file.write(parse_data(data_type, data_value))
                    
                    for row_index in range(0, data_row_count):

                        data_type = data_type_list[row_index]
                        data_value = excel_sheet_cols_data[data_id_col + 1 + col_index][data_id_row + row_index].value

                        export_file.write(",")
                        export_file.write("\"")
                        export_file.write(parse_data(data_type, data_value))
                        export_file.write("\"")
                    export_file.write("\n")

        export_file.close()
        print("Export " + file_id + " Success!")


def generate_code(file_id):
    excel_file = openpyxl.load_workbook(file_path + file_dict[file_id])
    file_config = config_json_data[file_id]
    sheet_list = file_config["sheet_list"]

    enable_server_code = False
    if "enable_server_code" in config_json_data \
            and config_json_data["enable_server_code"] == True \
            and "enable_server_code" in file_config \
            and file_config["enable_server_code"] == True:
        enable_server_code = True

    enable_client_code = False
    if "enable_client_code" in config_json_data \
            and config_json_data["enable_client_code"] == True \
            and "enable_client_code" in file_config \
            and file_config["enable_client_code"] == True:
        enable_client_code = True

    for sheet_info in sheet_list:

        excel_sheet = excel_file[sheet_info["sheet_name"]]

        data_id_row = sheet_info["data_id_row"] - 1
        data_id_col = sheet_info["data_id_col"] - 1

        data_id_value = excel_sheet.cell(row=data_id_row+1, column=data_id_col+1).value

        if data_id_value != "DataId":
            print("DataId row or col not valid!")
            return

        excel_sheet_rows_data = tuple(excel_sheet.rows)
        excel_sheet_cols_data = tuple(excel_sheet.columns)

        data_id_row_value_list = excel_sheet_rows_data[data_id_row]
        data_id_col_value_list = excel_sheet_cols_data[data_id_col]

        data_row_max = len(data_id_col_value_list)
        data_col_max = len(data_id_row_value_list)
        
        csv_head_list = []
        data_type_list = []

        if sheet_info["export_type"] == "Horizontal":

            for cell in excel_sheet_rows_data[data_id_row][data_id_col: data_col_max]:
                if cell.value in [None, ""]:
                    break
                csv_head_list.append(cell.value)
            
            for cell in excel_sheet_rows_data[data_id_row - 1][data_id_col: data_col_max]:
                if cell.value in [None, ""]:
                    break
                data_type_list.append(cell.value)

        elif sheet_info["export_type"] == "Vertical":

            csv_head_list = []
            for cell in excel_sheet_cols_data[data_id_col][data_id_row: data_row_max]:
                if cell.value in [None, ""]:
                    break
                csv_head_list.append(cell.value)

            for cell in excel_sheet_cols_data[data_id_col - 1][data_id_row: data_row_max]:
                            if cell.value in [None, ""]:
                                break
                            data_type_list.append(cell.value)

        if enable_server_code:
            with open(file_path + config_json_data['server_code_path'] + file_config["lower_case"] + config_json_data['server_file_suffix'] + config_json_data['server_file_extension'], 'w', encoding='utf-8') as server_code_file:

                server_code_file.write("#pragma once\n")
                server_code_file.write("\n")
                server_code_file.write("//Exported by Excel, please don't edit this file directly.\n")
                server_code_file.write("\n")
                server_code_file.write("#include \"" + config_json_data['type_def_file'] + "\"\n")
                server_code_file.write("#include \"" + config_json_data['enum_def_server_file'] + "\"\n")
                server_code_file.write("#include \"data_table_base.h\"\n")
                server_code_file.write("#include \"data_csv_parser.h\"\n")
                server_code_file.write("\n")
                server_code_file.write("namespace "+config_json_data["namespace"]+"\n")
                server_code_file.write("{\n")
                server_code_file.write("\tclass " + file_config["upper_case"] + " : public DataTableBase\n")
                server_code_file.write("\t{\n")
                server_code_file.write("\tpublic:\n")
                server_code_file.write("\t\t" + file_config["upper_case"] + "(std::string data_string)\n")
                server_code_file.write("\t\t{\n")
                server_code_file.write("\t\t\tDataCsvParser csv_parser(data_string);\n")

                for data_index in range(len(csv_head_list)):
                    server_code_file.write("\t\t\tcsv_parser.")
                    server_code_file.write(get_parse_function_name(data_type_list[data_index]))
                    server_code_file.write("(")
                    server_code_file.write(csv_head_list[data_index])
                    server_code_file.write(");\n")

                server_code_file.write("\t\t};\n")
                server_code_file.write("\n")

                for data_index in range(len(csv_head_list)):
                    server_code_file.write("\t\t")
                    server_code_file.write(get_server_data_type(data_type_list[data_index]))
                    server_code_file.write(" ")
                    server_code_file.write(csv_head_list[data_index])
                    server_code_file.write(";\n")

                server_code_file.write("\t};\n")
                server_code_file.write("}\n")

                server_code_file.close()
                print("Generate " + file_id + " Server Code Success!")

        if enable_client_code:
            with open(file_path + config_json_data['client_code_path'] + file_config["upper_case"] + config_json_data['client_file_suffix'] + config_json_data['client_file_extension'], 'w', encoding='utf-8') as client_code_file:

                client_code_file.write("#pragma once\n")
                client_code_file.write("\n")
                client_code_file.write("//Exported by Excel, please don't edit this file directly.\n")
                client_code_file.write("\n")
                client_code_file.write("#include \"" + config_json_data['type_def_file'] + "\"\n")
                client_code_file.write("#include \"" + config_json_data['enum_def_client_file'] + "\"\n")
                client_code_file.write("#include \"Engine/DataTable.h\"\n")
                client_code_file.write("#include \"" + file_config["upper_case"] + config_json_data['client_file_suffix'] + ".generated.h\"\n")
                client_code_file.write("\n")
                client_code_file.write("USTRUCT(BlueprintType)\n")
                client_code_file.write("struct F" + file_config["upper_case"] + " : public FTableRowBase\n")
                client_code_file.write("{\n")
                client_code_file.write("\tGENERATED_USTRUCT_BODY()\n")
                client_code_file.write("\n")

                for data_index in range(len(csv_head_list)):
                    client_code_file.write("\tUPROPERTY(EditAnywhere, BlueprintReadWrite, Category = F" + file_config["upper_case"] + ")\n")
                    client_code_file.write("\t\t")
                    client_code_file.write(get_client_data_type(data_type_list[data_index]))
                    client_code_file.write(" ")
                    client_code_file.write(csv_head_list[data_index])
                    client_code_file.write(";\n")

                client_code_file.write("};\n")

                client_code_file.close()
                print("Generate " + file_id + " Client Code Success!")

        break


def write_file_content(write_file_path, start_key, end_key, content):
    file_content = ""
    with open(write_file_path, 'r', encoding='utf-8') as target_file:
        file_content = target_file.read()
        start_index = file_content.find(start_key)
        if start_index == -1:
            return
        head_end_index = file_content.find(end_key)
        if head_end_index == -1:
            return
        start_content = ""
        end_content = ""
        start_content = file_content[:(start_index + len(start_key))]
        end_content = file_content[head_end_index:]
        file_content = start_content + content + end_content

    with open(write_file_path, 'w', encoding='utf-8') as target_file:
        target_file.write(file_content)


def register_datatable():
    server_header_content = "\n"
    server_register_content = "\n"
    server_declare_content = "\n"
    client_header_content = "\n"
    client_register_content = "\n"
    data_table_type = "\nenum EDataTableType\n{\n"
    data_table_escape = "\n"

    register_config = config_json_data["register"]

    for file_id in file_dict:

        file_config = config_json_data[file_id]

        enable_server_code = False
        if "enable_server_code" in config_json_data \
                and config_json_data["enable_server_code"] == True \
                and "enable_server_code" in file_config \
                and file_config["enable_server_code"] == True:
            enable_server_code = True

        enable_client_code = False
        if "enable_client_code" in config_json_data \
                and config_json_data["enable_client_code"] == True \
                and "enable_client_code" in file_config \
                and file_config["enable_client_code"] == True:
            enable_client_code = True

        if enable_server_code:
            server_header_content += "#include \"" + \
                file_config["lower_case"] + config_json_data['server_file_suffix'] + \
                config_json_data['server_file_extension'] + "\"\n"
            server_register_content += "\t\t\t, REGISTER_DATATABLE_SERVER(" + \
                file_config["upper_case"] + ")\n"
            server_declare_content += "\t\tDECLARE_DATATABLE_SERVER(" + \
                file_config["upper_case"] + ")\n"
        if enable_client_code:
            client_header_content += "#include \"" + \
                file_config["upper_case"] + config_json_data['client_file_suffix'] + \
                config_json_data['client_file_extension'] + "\"\n"
            client_register_content += "\tREGISTER_DATATABLE_CLIENT(" + \
                file_config["upper_case"] + ")\n"
        data_table_type += "\tEDataTableType_" + \
            file_config["upper_case"]+" = "+file_id+",\n"
        data_table_escape += "#define " + \
            file_config["upper_case"] + " F" + \
            file_config["upper_case"]+"\n"

    data_table_type += "\tEDataTableType_Max\n};\n"

    if enable_server_code:
        write_file_content(file_path + config_json_data['server_register_path'] + "data_table_manager.h",
                           register_config["server_header_start"], register_config["server_header_end"], server_header_content)
        write_file_content(file_path + config_json_data['server_register_path'] + "data_table_manager.h",
                           register_config["server_declare_start"], register_config["server_declare_end"], server_declare_content)
        write_file_content(file_path + config_json_data['server_register_path'] + "data_table_manager.cpp",
                           register_config["server_register_start"], register_config["server_register_end"], server_register_content)

    if enable_client_code:
        write_file_content(file_path + config_json_data['client_register_path'] + "DataTableManager.h",
                           register_config["client_header_start"], register_config["client_header_end"], client_header_content)
        write_file_content(file_path + config_json_data['client_register_path'] + "DataTableManager.cpp",
                           register_config["client_register_start"], register_config["client_register_end"], client_register_content)

    write_file_content(file_path + config_json_data['data_table_type_register_path'] + config_json_data["data_table_type_register_file"],
                       register_config["type_start"], register_config["type_end"], data_table_type)

    enable_escape = False
    if "enable_escape" in register_config \
            and register_config["enable_escape"] == True:
        enable_escape = True

    if enable_escape:
        write_file_content(file_path + config_json_data['data_table_type_register_path'] + config_json_data["data_table_type_register_file"],
                           register_config["escape_start"], register_config["escape_end"], data_table_escape)

    print("Register DataTable Success!")


def generate_enum():

    for file_id in file_dict:

        file_config = config_json_data[file_id]

        enable_server_code = False
        if "enable_server_code" in config_json_data \
                and config_json_data["enable_server_code"] == True \
                and "enable_server_code" in file_config \
                and file_config["enable_server_code"] == True:
            enable_server_code = True

        enable_client_code = False
        if "enable_client_code" in config_json_data \
                and config_json_data["enable_client_code"] == True \
                and "enable_client_code" in file_config \
                and file_config["enable_client_code"] == True:
            enable_client_code = True

        if enable_server_code:
            with open(file_path + config_json_data['enum_def_server_path'] + config_json_data['enum_def_server_file'], 'w', encoding='utf-8') as server_code_file:

                server_code_file.write("#pragma once\n")
                server_code_file.write("\n")
                server_code_file.write(
                    "//Exported by Json, please don't edit this file directly.\n")
                server_code_file.write("\n")
                server_code_file.write(
                    "#include \"" + config_json_data['type_def_file'] + "\"\n")
                server_code_file.write("\n")
                server_code_file.write(
                    "namespace "+config_json_data["namespace"]+"\n")
                server_code_file.write("{")

                for enum_type in enum_json_data:
                    server_code_file.write("\n")
                    server_code_file.write("\tenum "+enum_type)
                    if "#" in enum_json_data[enum_type]:
                        enum_comment = get_enum_type_comment(enum_type)
                        if enum_comment != "":
                            server_code_file.write(" //"+enum_comment)
                    server_code_file.write("\n")
                    server_code_file.write("\t{")

                    enum_data_list = list(enum_json_data[enum_type].keys())
                    for index in range(len(enum_data_list)):
                        enum_data = enum_data_list[index]
                        if enum_data == "#":
                            continue
                        server_code_file.write("\n\t\t"+enum_data)
                        server_code_file.write(" = ")
                        server_code_file.write(
                            str(get_enum_value(enum_type, enum_data)))
                        if(index != len(enum_data_list)-1):
                            server_code_file.write(",")
                        enum_comment = get_enum_comment(enum_type, enum_data)
                        if(enum_comment != ""):
                            server_code_file.write(" //"+enum_comment)
                    server_code_file.write("\n\t};\n")

                server_code_file.write("}\n")

                server_code_file.close()

        if enable_client_code:
            with open(file_path + config_json_data['enum_def_client_path'] + config_json_data['enum_def_client_file'], 'w', encoding='utf-8') as client_code_file:

                client_code_file.write("#pragma once\n")
                client_code_file.write("\n")
                client_code_file.write(
                    "//Exported by Json, please don't edit this file directly.\n")
                client_code_file.write("\n")
                client_code_file.write(
                    "#include \"" + config_json_data['type_def_file'] + "\"\n")

                for enum_type in enum_json_data:

                    client_code_file.write("\n")
                    client_code_file.write(
                        "UENUM(BlueprintType, Blueprintable)\n")
                    client_code_file.write("enum "+enum_type)
                    if "#" in enum_json_data[enum_type]:
                        enum_comment = get_enum_type_comment(enum_type)
                        if enum_comment != "":
                            client_code_file.write(" //"+enum_comment)
                    client_code_file.write("\n")
                    client_code_file.write("{")

                    enum_data_list = list(enum_json_data[enum_type].keys())
                    for index in range(len(enum_data_list)):
                        enum_data = enum_data_list[index]
                        if enum_data == "#":
                            continue
                        client_code_file.write("\n\t"+enum_data)
                        client_code_file.write(" = ")
                        client_code_file.write(
                            str(get_enum_value(enum_type, enum_data)))
                        if(index != len(enum_data_list)-1):
                            client_code_file.write(",")
                        enum_comment = get_enum_comment(enum_type, enum_data)
                        if(enum_comment != ""):
                            client_code_file.write(" //"+enum_comment)
                    client_code_file.write("\n};\n")

                client_code_file.close()
    print("Generate Enum Success!")


def process_file(file_id):
    print("1.export_data")
    print("2.generate_code & export_data")
    choose_idx = input("plase choose : ")
    if choose_idx == "1":
        export_data(file_id)
    elif choose_idx == "2":
        generate_code(file_id)
        export_data(file_id)
    else:
        print("not valid!")


def process_all_file():
    print("0.register_datatable")
    print("1.generate_enum")
    print("2.export_data")
    print("3.generate_code & export_data")
    choose_idx = input("please choose : ")

    if choose_idx == "0":
        register_datatable()
    elif choose_idx == "1":
        generate_enum()
    elif choose_idx == "2" or choose_idx == "3":
        for file_key in file_dict:
            if choose_idx == "3":
                generate_code(file_key)
            export_data(file_key)


def add_enum_value_data(key_a, key_b, val):
    if key_a in enum_value_data:
        enum_value_data[key_a].update({key_b: val})
    else:
        enum_value_data.update({key_a: {key_b: val}})


# 设置环境变量
file_path = os.path.dirname(os.path.abspath(sys.argv[0]))
os.chdir(file_path)
cwd = os.getcwd()

# 取数据配置
with open('../Config/Config.json', 'r', encoding='utf-8') as config_json_file:
    config_json_data = json.loads(config_json_file.read())
    config_json_file.close()

# 取数据配置
with open('../Config/Enum.json', 'r', encoding='utf-8') as enum_json_file:
    enum_json_data = json.loads(enum_json_file.read())
    enum_json_file.close()

enum_value_data = {}

for enum_type in enum_json_data:
    enum_list = list(enum_json_data[enum_type].keys())
    enum_index = 0
    for enum_data in enum_list:
        if enum_data == "#":
            continue
        enum_value = enum_json_data[enum_type][enum_data]
        if "Value" in enum_value:
            enum_index = int(enum_value["Value"])
        add_enum_value_data(enum_type, enum_data, enum_index)
        enum_index = enum_index + 1


if is_config_file_valid():
    # 取数据表
    file_dict = {}

    file_path = file_path + "\\..\\"
    file_list = os.listdir(file_path)
    for excel_file in file_list:
        if is_file_valid(excel_file[0:3]) and is_config_valid(excel_file[0:3]):
            file_dict[excel_file[0:3]] = excel_file
            # print(file_path + excel_file)

    # 用户选择
    print("Init Success!")
    print("-------------")
    for file_key in file_dict:
        print(file_key + "." + file_dict[file_key][3:])
    print("0.All")
    print("-------------")
    file_choose = input("Choose File : ")
    if file_choose == "0":
        process_all_file()
    elif not is_file_valid(file_choose):
        print("not valid!")
    else:
        process_file(file_choose)
